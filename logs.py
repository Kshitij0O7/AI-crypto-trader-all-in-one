"""
Logging module for trading bot - exports data to Excel files
"""
import os
from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")


def get_logs_directory() -> Path:
    """Get or create logs directory"""
    logs_dir = Path("logs")
    logs_dir.mkdir(exist_ok=True)
    return logs_dir


def get_log_filename(prefix: str = "trading_log") -> str:
    """Generate log filename with date"""
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{prefix}_{date_str}.xlsx"


def create_workbook_if_not_exists(filepath: Path) -> Workbook:
    """Create a new workbook or load existing one"""
    if filepath.exists():
        return load_workbook(filepath)
    else:
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb


def ensure_sheet_exists(wb: Workbook, sheet_name: str, headers: List[str]):
    """Ensure sheet exists with headers"""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
    else:
        ws = wb[sheet_name]
    return wb[sheet_name]


def log_open_positions(open_positions: List[Dict], market_data: Optional[Dict] = None, 
                       pnl_data: Optional[Dict] = None) -> Optional[str]:
    """
    Log open positions to Excel file
    
    Args:
        open_positions: List of open position dictionaries
        market_data: Optional market data for current prices
        pnl_data: Optional pre-calculated PnL data
    
    Returns:
        Filepath of created log file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        logs_dir = get_logs_directory()
        filepath = logs_dir / get_log_filename("open_positions")
        wb = create_workbook_if_not_exists(filepath)
        
        headers = [
            "Timestamp", "Position ID", "Market", "Asset ID", "Action", "Entry Price", 
            "Current Price", "Target Price", "Stop Loss", "Amount USD",
            "PnL USD", "PnL %", "Confidence", "Reasoning"
        ]
        
        ws = ensure_sheet_exists(wb, "Open Positions", headers)
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Get current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Log each position
        for position in open_positions:
            position_id = position.get('id', position.get('market', 'N/A'))
            market = position.get('market', position.get('token_out', 'N/A'))
            entry_price = position.get('entry_price', 0)
            current_price = entry_price  # Default to entry if not available
            
            # Try to get current price from pnl_data
            if pnl_data and pnl_data.get('positions'):
                for pnl_pos in pnl_data['positions']:
                    if pnl_pos.get('market') == market:
                        current_price = pnl_pos.get('current_price', entry_price)
                        break
            
            pnl_usd = 0
            pnl_pct = 0
            if pnl_data and pnl_data.get('positions'):
                for pnl_pos in pnl_data['positions']:
                    if pnl_pos.get('market') == market:
                        pnl_usd = pnl_pos.get('pnl_usd', 0)
                        pnl_pct = pnl_pos.get('pnl_pct', 0)
                        break
            
            row_data = [
                timestamp,
                position_id,
                market,
                position.get('asset_id', 'N/A'),  # Asset ID
                position.get('action', 'N/A'),
                entry_price,
                current_price,
                position.get('target_price', 0),
                position.get('stop_loss', 0),
                position.get('amount_usd', 0),
                pnl_usd,
                pnl_pct,
                position.get('confidence', 0),
                position.get('reasoning', '')[:100]  # Truncate long reasoning
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
        
        wb.save(filepath)
        return str(filepath)
        
    except Exception as e:
        print(f"❌ Error logging open positions: {e}")
        return None


def log_closed_positions(closed_positions: List[Dict]) -> Optional[str]:
    """
    Log closed positions to Excel file
    
    Args:
        closed_positions: List of closed position dictionaries
    
    Returns:
        Filepath of created log file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        logs_dir = get_logs_directory()
        filepath = logs_dir / get_log_filename("closed_positions")
        wb = create_workbook_if_not_exists(filepath)
        
        headers = [
            "Timestamp", "Position ID", "Market", "Asset ID", "Action", "Entry Price",
            "Exit Price", "Target Price", "Stop Loss", "Amount USD",
            "PnL USD", "PnL %", "Close Reason", "Confidence", "Reasoning"
        ]
        
        ws = ensure_sheet_exists(wb, "Closed Positions", headers)
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Log each closed position
        for position in closed_positions:
            timestamp = position.get('closed_at', position.get('timestamp', datetime.now().isoformat()))
            if isinstance(timestamp, str):
                try:
                    # Try to parse ISO format
                    dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            row_data = [
                timestamp,
                position.get('id', position.get('market', 'N/A')),
                position.get('market', position.get('token_out', 'N/A')),
                position.get('asset_id', 'N/A'),  # Asset ID
                position.get('action', 'N/A'),
                position.get('entry_price', 0),
                position.get('close_price', 0),
                position.get('target_price', 0),
                position.get('stop_loss', 0),
                position.get('amount_usd', 0),
                position.get('pnl_usd', 0),
                position.get('pnl_pct', 0),
                position.get('close_reason', 'N/A'),
                position.get('confidence', 0),
                position.get('reasoning', '')[:100]  # Truncate long reasoning
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
        
        wb.save(filepath)
        return str(filepath)
        
    except Exception as e:
        print(f"❌ Error logging closed positions: {e}")
        return None


def log_pnl_report(pnl_data: Dict, daily_pnl: float, timestamp: Optional[str] = None) -> Optional[str]:
    """
    Log PnL report to Excel file
    
    Args:
        pnl_data: Dictionary with PnL calculation results
        daily_pnl: Daily realized PnL
        timestamp: Optional timestamp string
    
    Returns:
        Filepath of created log file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        logs_dir = get_logs_directory()
        filepath = logs_dir / get_log_filename("pnl_reports")
        wb = create_workbook_if_not_exists(filepath)
        
        headers = [
            "Timestamp", "Total Open Positions", "Total Unrealized PnL USD",
            "Daily Realized PnL USD", "Total PnL USD"
        ]
        
        ws = ensure_sheet_exists(wb, "PnL Reports", headers)
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        total_pnl = daily_pnl + pnl_data.get('total_pnl', 0)
        
        row_data = [
            timestamp,
            pnl_data.get('position_count', 0),
            pnl_data.get('total_pnl', 0),
            daily_pnl,
            total_pnl
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)
        
        wb.save(filepath)
        return str(filepath)
        
    except Exception as e:
        print(f"❌ Error logging PnL report: {e}")
        return None


def log_signal_history(signal_history: List[Dict]) -> Optional[str]:
    """
    Log AI signal history to Excel file
    
    Args:
        signal_history: List of signal dictionaries
    
    Returns:
        Filepath of created log file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        logs_dir = get_logs_directory()
        filepath = logs_dir / get_log_filename("signal_history")
        wb = create_workbook_if_not_exists(filepath)
        
        headers = [
            "Timestamp", "Signal", "Outcome", "Market", "Asset ID", "Action", "Confidence"
        ]
        
        ws = ensure_sheet_exists(wb, "Signal History", headers)
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Log each signal
        for signal in signal_history:
            timestamp = signal.get('timestamp', datetime.now().isoformat())
            if isinstance(timestamp, str):
                try:
                    dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            signal_data = signal.get('signal', {})
            if isinstance(signal_data, dict):
                market = signal_data.get('market', 'N/A')
                action = signal_data.get('action', 'N/A')
                confidence = signal_data.get('confidence', 0)
                asset_id = signal_data.get('asset_id', 'N/A')  # Extract asset_id if available
            else:
                market = 'N/A'
                action = 'N/A'
                confidence = 0
                asset_id = 'N/A'
            
            row_data = [
                timestamp,
                str(signal_data)[:100],  # Truncate signal data
                signal.get('outcome', 'pending'),
                market,
                asset_id,  # Asset ID
                action,
                confidence
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
        
        wb.save(filepath)
        return str(filepath)
        
    except Exception as e:
        print(f"❌ Error logging signal history: {e}")
        return None


def log_trading_summary(open_positions: List[Dict], closed_positions: List[Dict],
                       daily_pnl: float, pnl_data: Optional[Dict] = None) -> Optional[str]:
    """
    Log complete trading summary to Excel file
    
    Args:
        open_positions: List of open positions
        closed_positions: List of closed positions
        daily_pnl: Daily realized PnL
        pnl_data: Optional PnL calculation data
    
    Returns:
        Filepath of created log file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        logs_dir = get_logs_directory()
        filepath = logs_dir / get_log_filename("trading_summary")
        wb = create_workbook_if_not_exists(filepath)
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Summary sheet
        summary_headers = ["Metric", "Value", "Timestamp"]
        ws_summary = ensure_sheet_exists(wb, "Summary", summary_headers)
        next_row = ws_summary.max_row + 1
        
        total_realized_pnl = sum(p.get('pnl_usd', 0) for p in closed_positions)
        total_unrealized_pnl = pnl_data.get('total_pnl', 0) if pnl_data else 0
        total_pnl = daily_pnl + total_unrealized_pnl
        
        summary_data = [
            ["Total Open Positions", len(open_positions), timestamp],
            ["Total Closed Positions", len(closed_positions), timestamp],
            ["Daily Realized PnL USD", daily_pnl, timestamp],
            ["Total Realized PnL USD", total_realized_pnl, timestamp],
            ["Total Unrealized PnL USD", total_unrealized_pnl, timestamp],
            ["Total PnL USD", total_pnl, timestamp],
        ]
        
        for row_data in summary_data:
            for col_idx, value in enumerate(row_data, start=1):
                ws_summary.cell(row=next_row, column=col_idx, value=value)
            next_row += 1
        
        # Also log open and closed positions in separate sheets
        if open_positions:
            log_open_positions(open_positions, pnl_data=pnl_data)
        
        if closed_positions:
            log_closed_positions(closed_positions)
        
        wb.save(filepath)
        return str(filepath)
        
    except Exception as e:
        print(f"❌ Error logging trading summary: {e}")
        return None
